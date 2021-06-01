# Program takes xlsx file type and does a price correction.
# Inserts price correct in the next row.
# Installs plot showing prices corrections for all product ID's.
# Saves new file over the old file automatically.

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def price_correction(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    correct_values = Reference(sheet, min_row=1, max_row=sheet.max_row,  min_col=4, max_col=4)
    old_values = Reference(sheet, min_row=1, max_row=sheet.max_row,  min_col=3, max_col=3)

    sheet['d1'] = 'Correct Prices'
    sheet['c1'] = 'Old Prices'

    chart = BarChart()
    chart.add_data(correct_values, titles_from_data=True)
    chart.add_data(old_values, titles_from_data=True)
    chart.type = "col"
    chart.style = 10
    chart.title = 'Comparative Price Changes'
    chart.y_axis.title = 'Price'
    chart.x_axis.title = 'Product ID'
    sheet.add_chart(chart, 'a7')

    wb.save(f'{filename}')


incorrect_file = input('Filename: ')
price_correction(incorrect_file)
